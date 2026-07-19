#!/usr/bin/env python3
"""Rebuild Phase 0 / canonical experimental-unit tables (reproducible)."""
from __future__ import annotations

import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "scfa_metabolomics"))
from obesity_equivalence_analysis import load_metadata  # noqa: E402

OUT = ROOT / "integrated" / "metadata"
OUT.mkdir(parents=True, exist_ok=True)


def main() -> None:
    meta = load_metadata()
    group_map = (
        meta.drop_duplicates("subject")
        .assign(donor_id=lambda d: d["subject"].astype(str))
        .set_index("donor_id")["group"]
        .to_dict()
    )

    scfa = pd.read_csv(
        ROOT
        / "scfa_metabolomics/data/removed_qcs_quant_results_20250722_PFBBr_PS2720_20250731.csv"
    )
    scfa["sample_id"] = scfa["sampleid"].astype(str).str.strip()
    scfa = scfa.loc[
        ~scfa["sample_id"].str.contains(
            "QC|Blank|blank|Standard|pool|Pool", case=False, na=False
        )
    ]
    parsed = scfa["sample_id"].str.extract(
        r"^(?P<root>\d+)(?P<aliq>[A-Za-z])\s+(?P<carb>[rs]\d|nc)\s+(?P<time>\d+)h$"
    )
    scfa = scfa.join(parsed).dropna(subset=["root"])
    scfa["donor_id"] = scfa["root"].astype(str)
    scfa["aliquot_id"] = scfa["root"] + scfa["aliq"].str.upper()
    scfa["carbohydrate"] = scfa["carb"].map(
        lambda x: "RDC"
        if str(x).startswith("r")
        else ("SDC" if str(x).startswith("s") else "No Added Carb")
    )
    scfa["well_repeat"] = (
        scfa["carb"].str.extract(r"(\d+)").fillna("1").astype(int)
    )
    scfa["timepoint"] = scfa["time"].astype(int).astype(str) + "H"
    scfa["well_id"] = (
        scfa["aliquot_id"]
        + "::"
        + scfa["carbohydrate"]
        + "::"
        + scfa["well_repeat"].astype(str)
    )
    scfa["group"] = scfa["donor_id"].map(group_map)
    scfa["group_label"] = scfa["group"].map(
        {"control": "healthy-weight", "case": "obesity"}
    )
    scfa["assay"] = "scfa"
    canon_scfa = scfa[
        [
            "sample_id",
            "donor_id",
            "aliquot_id",
            "carbohydrate",
            "well_repeat",
            "well_id",
            "timepoint",
            "group",
            "group_label",
            "assay",
        ]
    ].drop_duplicates()

    wb = load_workbook(
        ROOT / "scfa_metabolomics/metadata/SEED LAB HMM Submission Sheet 202405.xlsx",
        data_only=True,
    )
    ids = []
    for row in wb.active.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str) and re.match(r"^\d+[A-Za-z]\s", cell.strip()):
                ids.append(cell.strip())
    planned_ab = sorted(
        {re.match(r"^(\d+[A-Za-z])", x).group(1).upper() for x in ids}
    )
    planned_donors = sorted(
        {re.match(r"^(\d+)", a).group(1) for a in planned_ab}, key=int
    )
    analyzed_ab = sorted(canon_scfa["aliquot_id"].unique())
    analyzed_donors = sorted(canon_scfa["donor_id"].unique(), key=int)

    micro = pd.read_csv(
        ROOT
        / "microbiome_analysis/zr24558.16S_250813.zymo/00...AllSamples.Bac16Sv34"
        / "Sample_Information/sample.metadata.csv",
        sep="\t",
    )
    micro["customer_label"] = micro["customer_label"].astype(str).str.strip()
    mparse = micro["customer_label"].str.extract(
        r"^(?P<aliquot_id>\d+[A-Za-z])\.(?P<rep>[A-Za-z0-9]+)\.(?P<timepoint>\d+H)$"
    )
    micro = micro.join(mparse).dropna(subset=["aliquot_id"])
    micro["aliquot_id"] = micro["aliquot_id"].str.upper()
    micro["donor_id"] = micro["aliquot_id"].str.extract(r"^(\d+)")[0].astype(str)
    rep = micro["rep"].astype(str).str.upper()
    micro["carbohydrate"] = [
        "RDC"
        if r.startswith("R")
        else ("SDC" if r.startswith("S") else ("No Added Carb" if r.startswith("N") else None))
        for r in rep
    ]
    micro["well_repeat"] = rep.str.extract(r"(\d+)").fillna("1").astype(int)
    micro["well_id"] = (
        micro["aliquot_id"]
        + "::"
        + micro["carbohydrate"].astype(str)
        + "::"
        + micro["well_repeat"].astype(str)
    )
    micro["group"] = micro["donor_id"].map(group_map)
    micro["group_label"] = micro["group"].map(
        {"control": "healthy-weight", "case": "obesity"}
    )
    micro["sample_id"] = micro["sample_id"].astype(str)
    micro["assay"] = "16S"
    canon_micro = micro[
        [
            "sample_id",
            "donor_id",
            "aliquot_id",
            "carbohydrate",
            "well_repeat",
            "well_id",
            "timepoint",
            "group",
            "group_label",
            "assay",
        ]
    ].drop_duplicates()

    planned = pd.DataFrame(
        {
            "donor_id": [re.match(r"^(\d+)", a).group(1) for a in planned_ab],
            "aliquot_id": planned_ab,
            "in_scfa_export": [a in set(analyzed_ab) for a in planned_ab],
            "in_microbiome_metadata": [
                a in set(canon_micro["aliquot_id"]) for a in planned_ab
            ],
        }
    )
    planned["group"] = planned["donor_id"].map(group_map)
    planned["group_label"] = planned["group"].map(
        {"control": "healthy-weight", "case": "obesity"}
    )

    design_rows = []
    for aid in analyzed_ab:
        donor = re.match(r"^(\d+)", aid).group(1)
        ali = aid[-1]
        for carb, wells in [
            ("No Added Carb", [1]),
            ("RDC", [1, 2]),
            ("SDC", [1, 2]),
        ]:
            for w in wells:
                for tp in ("0H", "48H"):
                    if carb == "No Added Carb":
                        token = "nc"
                    elif carb == "RDC":
                        token = f"r{w}"
                    else:
                        token = f"s{w}"
                    expected = f"{donor}{ali.lower()} {token} {tp[:-1].lower()}h"
                    present = canon_scfa["sample_id"].str.lower().eq(expected).any()
                    design_rows.append(
                        {
                            "donor_id": donor,
                            "aliquot_id": aid,
                            "carbohydrate": carb,
                            "well_repeat": w,
                            "timepoint": tp,
                            "expected_sample_id": expected,
                            "present_in_scfa_export": bool(present),
                        }
                    )
    design = pd.DataFrame(design_rows)
    missing_samples = design.loc[~design["present_in_scfa_export"]].copy()
    missing_donors = sorted(set(planned_donors) - set(analyzed_donors), key=int)

    summary = pd.DataFrame(
        [
            {"metric": "planned_donors", "value": len(planned_donors)},
            {"metric": "planned_aliquots", "value": len(planned_ab)},
            {"metric": "scfa_analyzed_donors", "value": len(analyzed_donors)},
            {"metric": "scfa_analyzed_aliquots", "value": len(analyzed_ab)},
            {"metric": "scfa_missing_planned_donors", "value": len(missing_donors)},
            {
                "metric": "scfa_missing_planned_aliquots",
                "value": int((~planned["in_scfa_export"]).sum()),
            },
            {
                "metric": "scfa_missing_design_samples",
                "value": int((~design["present_in_scfa_export"]).sum()),
            },
            {
                "metric": "microbiome_aliquots",
                "value": int(canon_micro["aliquot_id"].nunique()),
            },
            {
                "metric": "microbiome_donors",
                "value": int(canon_micro["donor_id"].nunique()),
            },
            {"metric": "microbiome_rows", "value": len(canon_micro)},
            {
                "metric": "donors_in_both_scfa_and_16s",
                "value": len(
                    set(canon_scfa["donor_id"]) & set(canon_micro["donor_id"])
                ),
            },
            {"metric": "baseline_git_sha", "value": "67bd320"},
        ]
    )

    canon_scfa.to_csv(OUT / "canonical_experimental_units_scfa.csv", index=False)
    canon_micro.to_csv(OUT / "canonical_experimental_units_16s.csv", index=False)
    pd.concat([canon_scfa, canon_micro], ignore_index=True).to_csv(
        OUT / "canonical_experimental_units.csv", index=False
    )
    summary.to_csv(OUT / "phase0_unit_inventory_summary.csv", index=False)
    planned.to_csv(OUT / "phase0_planned_aliquots.csv", index=False)
    (
        canon_scfa.groupby(["donor_id", "aliquot_id"], as_index=False)
        .agg(n_samples=("sample_id", "nunique"), group=("group", "first"))
        .to_csv(OUT / "phase0_scfa_analyzed_aliquots.csv", index=False)
    )
    missing_samples.to_csv(OUT / "phase0_scfa_missing_design_samples.csv", index=False)
    pd.DataFrame({"donor_id": missing_donors}).to_csv(
        OUT / "phase0_scfa_missing_planned_donors.csv", index=False
    )
    print(summary.to_string(index=False))
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
